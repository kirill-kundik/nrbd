import collections
from typing import List, Union, Tuple

import openpyxl
import psycopg2
from openpyxl.worksheet.worksheet import Worksheet

import database


class XlsxWrapperError(Exception):
    pass


class XlsxWrapperInsertionError(XlsxWrapperError):
    pass


class XlsxWrapper:
    DIST_VALUE_NAMES: dict = {
        'mean': 'Мат. сподів.',
        'std': 'Середнє квадратичне відхилення',
        'mode': 'Мода',
        'min': 'Мін.',
        'max': 'Макс.',
        'coeff': 'Коефіцієнт варіації'
    }

    def __init__(self, filename: str):
        self._filename: str = filename
        self._workbook: openpyxl.Workbook = openpyxl.Workbook()
        self._rows: collections.defaultdict = collections.defaultdict(lambda: 1)

    def save(self) -> None:
        if 'Sheet' in self._workbook and not bool(self._workbook['Sheet']._cells):
            del self._workbook['Sheet']

        self._workbook.save(self._filename)

    def sheet_names(self) -> List[str]:
        return self._workbook.sheetnames

    def get_sheet(self, sheet_name: str) -> Union[Worksheet, None]:
        """
        :param sheet_name: title of the sheet where distances should be inserted
        :return: openxypl.Worksheet or None
        """
        return self._workbook[sheet_name] if sheet_name in self._workbook else None

    def get_or_create_sheet(self, sheet_name: str) -> Worksheet:
        if sheet_name in self._workbook:
            return self._workbook[sheet_name]
        else:
            return self._workbook.create_sheet(sheet_name)

    def insert_distances(self, sheet_name: str, distances: list) -> None:
        """
        insert_distances('ALL', [1, 2, 3])
        :exception: XlsxWrapperInsertionError
        :param sheet_name: title of the sheet where distances should be inserted
        :param distances: array object of printable values that will be inserted as distances
        :return:
        """
        if sheet_name in self._rows and self._rows[sheet_name] > 1:
            raise XlsxWrapperInsertionError(
                f'cannot insert distances on the first position: sheet \'{sheet_name}\' was edited previously'
            )

        sheet: Worksheet = self.get_or_create_sheet(sheet_name)
        self._insert_row(sheet, ['Відстань', *distances])

    def insert_distribution(self, sheet_name: str, distribution: dict) -> None:
        """
        insert_distribution(
            'ALL', {
                'dist_1': [1, 2, 3],
                'dist_2': [2, 3, 4],
                'values': {
                    'mean': 1.0,
                    'std': 1.0,
                    'mode': 1.0,
                    'min': 1.0,
                    'max': 1.0,
                    'coeff': 1.0
                }
            }
        )
        :exception: XlsxWrapperInsertionError
        :param sheet_name: title of the sheet where distributions should be inserted
        :param distribution: dict object that should store keys as
        name of inserted distributions and statistical data under 'values' key (REQUIRED!).
        :return:
        """
        if sheet_name not in self._rows or self._rows[sheet_name] == 1:
            raise XlsxWrapperInsertionError(
                f'cannot insert distribution on the first position: sheet \'{sheet_name}\' was not edited previously'
            )
        dist_values: dict = distribution.pop('values', None)
        if dist_values is None:
            raise XlsxWrapperInsertionError(f'cannot insert distribution: values was not provided')

        sheet: Worksheet = self.get_or_create_sheet(sheet_name)

        for dist_name, dist in distribution.items():
            self._insert_row(sheet, self._transform_distribution(dist_name, dist))

        for row in self._transform_distribution_values(dist_values):
            if row:
                self._insert_row(sheet, row, col=2)

        self._insert_blank_row(sheet)

    def insert_wild_type(
            self,
            sheet_name: str,
            wild_type: str,
            wild_type_poly_rcrs: float,
            wild_type_poly_rsrs: float,
            population_poly_rcrs: float,
            population_poly_rsrs: float
    ):
        """
        insert_wild_type('ALL', 'AAAAA', 10, 15, 1, 0)
        :exception: XlsxWrapperInsertionError
        :param sheet_name: title of the sheet where distributions should be inserted
        :param wild_type: wild type fasta code
        :param wild_type_poly_rcrs: Кількість поліморфізмів у дикого типу відносно базової rCRS
        :param wild_type_poly_rsrs: Кількість поліморфізмів у дикого типу відносно базової RSRS
        :param population_poly_rcrs: Кількість поліморфізмів у популяції відносно базової rCRS
        :param population_poly_rsrs: Кількість поліморфізмів у популяції відносно базової RSRS
        :return:
        """
        if sheet_name not in self._rows or self._rows[sheet_name] == 1:
            raise XlsxWrapperInsertionError(
                f'cannot insert wild type on the first position: sheet \'{sheet_name}\' was not edited previously'
            )

        sheet: Worksheet = self.get_or_create_sheet(sheet_name)

        self._insert_row(sheet, ['Рядочок дикого типу', wild_type], col=2)
        self._insert_row(
            sheet, ['Кількість поліморфізмів у дикого типу відносно базової rCRS', wild_type_poly_rcrs], col=2
        )
        self._insert_row(
            sheet, ['Кількість поліморфізмів у дикого типу відносно базової RSRS', wild_type_poly_rsrs], col=2
        )
        self._insert_row(
            sheet, ['Кількість поліморфізмів у популяції відносно базової rCRS', population_poly_rcrs], col=2
        )
        self._insert_row(
            sheet, ['Кількість поліморфізмів у популяції відносно базової RSRS', population_poly_rsrs], col=2
        )
        self._insert_blank_row(sheet)

    def _insert_blank_row(self, sheet: Worksheet) -> None:
        self._rows[sheet.title] += 1

    def _insert_row(self, sheet: Worksheet, values: list, row: int = None, col: int = None) -> bool:
        if row is None:
            write_row: int = self._rows[sheet.title]
            self._rows[sheet.title] += 1
        else:
            write_row: int = row

        write_col: int = 1 if col is None else col

        for value in values:
            sheet.cell(row=write_row, column=write_col, value=value)
            write_col += 1

        return True

    @staticmethod
    def _transform_distribution(dist_name: str, dist: list) -> list:
        return [dist_name, *dist]

    @staticmethod
    def _transform_distribution_values(dist_values: dict) -> Tuple[List[str], List]:
        names: List[str] = []
        values: list = []

        for key in XlsxWrapper.DIST_VALUE_NAMES:
            if key in dist_values:
                names.append(XlsxWrapper.DIST_VALUE_NAMES[key])
                values.append(dist_values[key])

        return names, values


def create_sheet(wrapper, db, region):
    # First row of the document: distances
    print(region)
    print("__________________________________")
    dist_range = range(0, 20)
    wrapper.insert_distances(region, dist_range)
    # №№2-5,7-10,12-15,17-20 rows in the docs
    db.calculate_wild(region)
    conn.commit()
    wild_type = db.select("public.sequence", [f"name='WILD_TYPE_{region}'"], first_=True)
    print(wild_type)

    for base_name in ['EVA', 'ANDREWS', f'WILD_TYPE_{region}']:
        res = db.distribution(base_name, region)
        line_1 = [0 for i in dist_range]
        line_2 = [0 for i in dist_range]
        for i in dist_range:
            for j in res:
                if j["diff_num"] == i:
                    line_1[i] = j["frequency"]
                    line_2[i] = j["p"]

        math_expectation = db.math_expectation(base_name, region)[0]['math_expectation']
        std = db.std(base_name, region)[0]['standart_dev']
        mode = db.mode(base_name, region)[0]['diff_num']
        min_value = db.min_value(base_name, region)[0]['min']
        max_value = db.max_value(base_name, region)[0]['max']
        coeff = db.coeff(base_name, region)[0]['koef']
        wrapper.insert_distribution(
            region, {f'Розподіл відносно {base_name}': line_1, f'Розподіл відносно {base_name} (частка)': line_2, 'values': {
                'mean': math_expectation,
                'std': std,
                'mode': mode,
                'min': min_value,
                'max': max_value,
                'coeff': coeff}}
        )

    res = db.distribution_each_to_each(region)
    line_1 = [0 for i in dist_range]
    line_2 = [0 for i in dist_range]
    for i in dist_range:
        for j in res:
            if j["diff_num"] == i:
                line_1[i] = j["frequency"]
                line_2[i] = j["p"]

    math_expectation = db.math_expectation_each_to_each(region)[0]['math_expectation']
    print(math_expectation)
    std = db.std_each_to_each(region)[0]['standart_dev']
    print(std)
    mode = db.mode_each_to_each(region)[0]['diff_num']
    print(mode)
    min_value = db.min_value_each_to_each(region)[0]['min']
    print(min_value)
    max_value = db.max_value_each_to_each(region)[0]['max']
    print(max_value)
    coeff = db.coeff_each_to_each(region)[0]['koef']
    print(coeff)
    wrapper.insert_distribution(
        region, {'Розподіл кожен з кожним': line_1, 'Розподіл кожен з кожним (частка)': line_2, 'values': {
            'mean': math_expectation,
                'std': std,
                'mode': mode,
                'min': min_value,
                'max': max_value,
                'coeff': coeff}}
    )
    # The last rows of the sheet with wild type and population statistics
    polim_ANDREWS_WILD = db.diff_between_base_and_wild('ANDREWS', f'WILD_TYPE_{region}')
    if polim_ANDREWS_WILD:
        pol1 = polim_ANDREWS_WILD[0][0]
    else:
        pol1 = 0
    polim_EVA_WILD = db.diff_between_base_and_wild('EVA', f'WILD_TYPE_{region}')
    if polim_EVA_WILD:
        pol2 = polim_EVA_WILD[0][0]
    else:
        pol2 = 0
    polim_EVA = db.polim('EVA', region)
    if polim_EVA:
        pol3 = polim_EVA[0][0]
    else:
        pol3 = 0
    polim_Andrews = db.polim('ANDREWS', region)
    if polim_Andrews:
        pol4 = polim_Andrews[0][0]
    else:
        pol4 = 0
    wrapper.insert_wild_type(region, wild_type[3], pol1, pol2, pol3, pol4)


if __name__ == '__main__':
    conn = psycopg2.connect("dbname='nrbd' user='postgres' host='localhost' password='gulayeva'")
    # conn = psycopg2.connect("dbname='nrbd' user='postgres' host='localhost' password='root'")
    # conn.set_session(autocommit=True) # enabling autocommit
    db = database.Database(conn)
    wrapper = XlsxWrapper('final.xlsx')
    regions = ['ALL', 'IF', 'BK', 'BG', 'ST', 'CH', 'KHM']
    for region in regions:
        create_sheet(wrapper, db, region)

    wrapper.save()
